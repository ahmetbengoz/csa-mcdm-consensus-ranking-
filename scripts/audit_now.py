from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def safe_read_excel(xlsx_path: Path, sheet_name: str) -> pd.DataFrame:
    try:
        return pd.read_excel(xlsx_path, sheet_name=sheet_name)
    except Exception:
        return pd.DataFrame()


def get_actual_sheet_names(xlsx_path: Path) -> list[str]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    return wb.sheetnames


def make_sheet_inventory(xlsx_path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows.append(
            {
                "sheet_name": sheet_name,
                "n_rows": ws.max_row,
                "n_cols": ws.max_column,
            }
        )
    return pd.DataFrame(rows)


def make_range_check(xlsx_path: Path, sheet_names: list[str]) -> pd.DataFrame:
    rows = []

    for sheet_name in sheet_names:
        df = safe_read_excel(xlsx_path, sheet_name)

        if df is None or len(df.columns) == 0:
            rows.append(
                {
                    "sheet_name": sheet_name,
                    "column_name": "__EMPTY_SHEET__",
                    "dtype_detected": "empty",
                    "n_rows": 0,
                    "n_nonnull": 0,
                    "n_missing": 0,
                    "n_unique": 0,
                    "min_value": np.nan,
                    "max_value": np.nan,
                    "mean_value": np.nan,
                }
            )
            continue

        for col in df.columns:
            s = df[col]
            numeric = pd.to_numeric(s, errors="coerce")

            rows.append(
                {
                    "sheet_name": sheet_name,
                    "column_name": str(col),
                    "dtype_detected": str(s.dtype),
                    "n_rows": len(s),
                    "n_nonnull": int(s.notna().sum()),
                    "n_missing": int(s.isna().sum()),
                    "n_unique": int(s.nunique(dropna=True)),
                    "min_value": float(numeric.min()) if numeric.notna().any() else np.nan,
                    "max_value": float(numeric.max()) if numeric.notna().any() else np.nan,
                    "mean_value": float(numeric.mean()) if numeric.notna().any() else np.nan,
                }
            )

    return pd.DataFrame(rows)


def audit_criteria_weights(xlsx_path: Path) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    df = safe_read_excel(xlsx_path, "Criteria_Weights").copy()

    if df.empty:
        raise ValueError("Criteria_Weights sheet could not be read or is empty.")

    required = {"SubCode", "MainCode", "PropOfMain", "AbsWeight"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Criteria_Weights sheet missing columns: {missing}")

    df["PropOfMain"] = pd.to_numeric(df["PropOfMain"], errors="coerce")
    df["AbsWeight"] = pd.to_numeric(df["AbsWeight"], errors="coerce")

    main_summary = (
        df.groupby("MainCode", as_index=False)["AbsWeight"]
        .sum()
        .rename(columns={"AbsWeight": "main_weight_reconstructed"})
        .sort_values("MainCode")
        .reset_index(drop=True)
    )

    within_main = (
        df.groupby("MainCode", as_index=False)["PropOfMain"]
        .sum()
        .rename(columns={"PropOfMain": "prop_of_main_sum"})
        .sort_values("MainCode")
        .reset_index(drop=True)
    )
    within_main["prop_sum_ok"] = np.isclose(within_main["prop_of_main_sum"], 1.0, atol=1e-8)

    main_summary = main_summary.merge(within_main, on="MainCode", how="left")

    summary = {
        "n_subcriteria": int(df["SubCode"].nunique()),
        "n_main_criteria": int(df["MainCode"].nunique()),
        "abs_weight_total": float(df["AbsWeight"].sum()),
        "abs_weight_total_ok": bool(np.isclose(df["AbsWeight"].sum(), 1.0, atol=1e-8)),
    }

    return df, main_summary, summary


def audit_alternative_scores(xlsx_path: Path) -> tuple[pd.DataFrame, dict]:
    df = safe_read_excel(xlsx_path, "Alternative_Scores").copy()

    if df.empty:
        raise ValueError("Alternative_Scores sheet could not be read or is empty.")

    first_col = df.columns[0]
    df = df.rename(columns={first_col: "alternative_id"})

    sub_cols = [c for c in df.columns if c != "alternative_id"]

    long_df = df.melt(
        id_vars=["alternative_id"],
        value_vars=sub_cols,
        var_name="sub_code",
        value_name="raw_score",
    )
    long_df["raw_score"] = pd.to_numeric(long_df["raw_score"], errors="coerce")

    summary = {
        "n_alternatives": int(df["alternative_id"].nunique()),
        "n_subcriteria_columns": int(len(sub_cols)),
        "score_min": float(long_df["raw_score"].min()),
        "score_max": float(long_df["raw_score"].max()),
        "score_missing": int(long_df["raw_score"].isna().sum()),
    }

    return long_df, summary


def audit_method_reliability(xlsx_path: Path) -> pd.DataFrame:
    df = safe_read_excel(xlsx_path, "Method_Reliability").copy()
    if not df.empty and "RelWeight" in df.columns:
        df["RelWeight"] = pd.to_numeric(df["RelWeight"], errors="coerce")
        df["weight_sum_total"] = df["RelWeight"].sum()
        df["weight_sum_ok"] = np.isclose(df["RelWeight"].sum(), 1.0, atol=1e-8)
    return df


def audit_rmse(xlsx_path: Path) -> pd.DataFrame:
    df = safe_read_excel(xlsx_path, "RMSE").copy()
    if not df.empty and "RMSE" in df.columns:
        df["RMSE"] = pd.to_numeric(df["RMSE"], errors="coerce")
    return df


def audit_spearman(xlsx_path: Path) -> pd.DataFrame:
    df = safe_read_excel(xlsx_path, "Spearman_vs_CSA").copy()
    return df


def audit_csi(xlsx_path: Path) -> pd.DataFrame:
    df = safe_read_excel(xlsx_path, "CSI").copy()
    return df


def write_summary(
    out_dir: Path,
    criteria_summary: dict,
    alt_summary: dict,
    reliability_df: pd.DataFrame,
    rmse_df: pd.DataFrame,
    spearman_df: pd.DataFrame,
    csi_df: pd.DataFrame,
) -> None:
    lines = []
    lines.append("AUDIT SUMMARY")
    lines.append("=" * 50)
    lines.append(f"n_main_criteria: {criteria_summary['n_main_criteria']}")
    lines.append(f"n_subcriteria: {criteria_summary['n_subcriteria']}")
    lines.append(f"abs_weight_total: {criteria_summary['abs_weight_total']}")
    lines.append(f"abs_weight_total_ok: {criteria_summary['abs_weight_total_ok']}")
    lines.append(f"n_alternatives: {alt_summary['n_alternatives']}")
    lines.append(f"n_subcriteria_columns: {alt_summary['n_subcriteria_columns']}")
    lines.append(f"score_min: {alt_summary['score_min']}")
    lines.append(f"score_max: {alt_summary['score_max']}")
    lines.append(f"score_missing: {alt_summary['score_missing']}")

    if not reliability_df.empty and "RelWeight" in reliability_df.columns:
        lines.append(f"method_reliability_sum: {reliability_df['RelWeight'].sum()}")

    if not rmse_df.empty:
        lines.append(f"rmse_rows: {len(rmse_df)}")

    if not spearman_df.empty:
        lines.append(f"spearman_rows: {len(spearman_df)}")

    if not csi_df.empty:
        numeric_values = pd.to_numeric(csi_df.stack(), errors='coerce').dropna().tolist()
        if numeric_values:
            lines.append(f"first_csi_numeric_value: {numeric_values[0]}")

    (out_dir / "audit_summary.txt").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--outdir", default="outputs\\audit_now")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).resolve()
    out_dir = Path(args.outdir).resolve()
    ensure_dir(out_dir)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")

    sheet_names = get_actual_sheet_names(xlsx_path)

    inventory = make_sheet_inventory(xlsx_path)
    inventory.to_csv(out_dir / "sheet_inventory.csv", index=False)

    range_check = make_range_check(xlsx_path, sheet_names)
    range_check.to_csv(out_dir / "range_check.csv", index=False)

    criteria_raw, criteria_main, criteria_summary = audit_criteria_weights(xlsx_path)
    criteria_raw.to_csv(out_dir / "criteria_weights_raw.csv", index=False)
    criteria_main.to_csv(out_dir / "criteria_main_weights.csv", index=False)

    alternative_scores_long, alt_summary = audit_alternative_scores(xlsx_path)
    alternative_scores_long.to_csv(out_dir / "alternative_scores_long.csv", index=False)

    reliability_df = audit_method_reliability(xlsx_path)
    reliability_df.to_csv(out_dir / "method_reliability.csv", index=False)

    rmse_df = audit_rmse(xlsx_path)
    rmse_df.to_csv(out_dir / "rmse.csv", index=False)

    spearman_df = audit_spearman(xlsx_path)
    spearman_df.to_csv(out_dir / "spearman_vs_csa.csv", index=False)

    csi_df = audit_csi(xlsx_path)
    csi_df.to_csv(out_dir / "csi.csv", index=False)

    write_summary(
        out_dir=out_dir,
        criteria_summary=criteria_summary,
        alt_summary=alt_summary,
        reliability_df=reliability_df,
        rmse_df=rmse_df,
        spearman_df=spearman_df,
        csi_df=csi_df,
    )

    print(f"[OK] Audit outputs written to: {out_dir}")


if __name__ == "__main__":
    main()